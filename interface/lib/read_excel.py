import xlrd
from openpyxl import load_workbook


def excel_to_list(data_file, sheet):
    data_list = []  # 新建个空列表，来乘装所有的数据
    wb = xlrd.open_workbook(data_file)  # 打开excel
    sh = wb.sheet_by_name(sheet)  # 获取工作簿
    header = sh.row_values(0)  # 获取标题行数据
    for i in range(1, sh.nrows):  # 跳过标题行，从第二行开始取数据
        d = dict(zip(header, sh.row_values(i)))  # 将标题和每行数据组装成字典
        data_list.append(d)
    return data_list  # 列表嵌套字典格式，每个元素是一个字典


def get_test_data(data_list, case_name):
    for case_data in data_list:
        if case_name == case_data['case_name']:  # 如果字典数据中case_name与参数一致
            return case_data
            # 如果查询不到会返回None


def openpyxl_test(data_file,data):
    print(type(data))
    excel = load_workbook(data_file)
    table = excel["TestUserLogin"]
    print(table)
    # 获取行数，列数
    # 获取单元格的值
   # data = table.cell(row=1, column=7).value
    sheet = excel.active
    maxr = sheet.max_row  # 列
    for i in range(2, maxr + 1):
         if data==0:
            sheet.cell(row=i, column=7, value="pass")
            excel.save("D:/腾讯课堂/interfaceAutomation/interface/data/test_user_data7.xlsx")
         elif data==1:
            sheet.cell(row=i, column=7, value="fail")
            excel.save("D:/腾讯课堂/interfaceAutomation/interface/data/test_user_data7.xlsx")


if __name__ == '__main__':  # 测试一下自己的代码
    #
    # data_list = excel_to_list(r"D:\腾讯课堂\interfaceAutomation\test_user_data.xlsx",
    #                           "TestUserLogin")  # 读取excel，TestUserLogin工作簿的所有数据
    # cases = {"test_user_login_normal", "test_user_login_password_wrong"}
    # for case in cases:
    #     case_data = get_test_data(data_list, case)  # 查找用例'test_user_login_normal'的数据
    #     print(case_data)
    openpyxl_test(r"D:\腾讯课堂\interfaceAutomation\interface\data\test_user_data.xlsx",0)
